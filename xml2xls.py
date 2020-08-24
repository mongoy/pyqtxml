import os
import fnmatch
import datetime
import xlsxwriter
import xlwt
import sys
import sqlite3
import xml.etree.cElementTree as ET
import openpyxl as opx

from PyQt5 import QtSql
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import (QWidget, QLabel, QApplication, QPushButton, QMainWindow)


BASE_DIR = os.getcwd()
global comp_naim, num_files


class MyParser(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Инициализация программы
        self.quit_btn = QPushButton('Выход', self)
        self.ok_btn = QPushButton('ОК', self)
        self.font = QFont()
        self.file_lbl = QLabel(f'Файлы в директории {os.path.join(BASE_DIR,"xml")} ', self)
        self.initUI()

    def initUI(self):
        # Меню
        menu = self.menuBar()

        # Настройки шрифта текста.
        self.font.setFamily("Arial")
        self.font.setPointSize(12)

        # Создадим надпись
        self.file_lbl.setFont(self.font)
        self.file_lbl.move(10, 10)
        self.file_lbl.resize(300, 50)

        # Создаем кнопку "OK".
        self.ok_btn.resize(100, 25)
        self.ok_btn.move(270, 140)

        # Создаем кнопку "Выход".
        self.quit_btn.resize(100, 25)
        self.quit_btn.move(270, 170)

        # Выполнить парсинг файлов
        self.ok_btn.clicked.connect(self.makeRequest)

        # Выход
        self.quit_btn.clicked.connect(QApplication.instance().quit)

        self.setFixedSize(400, 200)
        self.setWindowTitle('Парсер XML HWINFO')
        self.show()

    def makeRequest(self):
        """
        После нажатия на "ОК" выполняется запрос к API с выбранными данными.
        """
        # print(xml_to_xlsx())
        self.file_lbl.setText(f'Обработано - {xml_to_xlsx()} файлов')
        self.ok_btn.setEnabled(False)


def xml_to_xlsx():
    """
    Прсер xml-файла (программа HWINFO)
    """

    dir_xml = os.path.join(BASE_DIR, 'xml')
    files = fnmatch.filter(os.listdir(dir_xml), "*.XML")  # только XML
    # открываем новый файл на запись
    wb = opx.load_workbook(os.path.join(BASE_DIR, "blank\\zhelezo.xlsx"))
    # Страница
    ws = wb.active  # wb['Sheet1']  #
    # Строка, столбец
    num_row = 2
    num_col = 2

    for num_files, file in enumerate(files):
        file_xml = os.path.join(dir_xml, file)
        tree = ET.parse(file_xml)
        root = tree.getroot()  # корневой элемент
        # print(file_xml)
        # Системник
        comp_par = ['Computer Name', 'Operating System']
        # print('COMPUTER')
        parametr = []
        for elements in root.iter('COMPUTER'):
            for child in elements:
                if child.tag == 'NodeName':  # Имя компа
                    comp_naim = child.text
                    parametr = [comp_naim, 'Computer']
                    # parametr.append(comp_naim)
                    # parametr.append('Computer')
                if child.tag == 'Property':  # Параметры компа
                    if child[0].text in comp_par:
                        print(f'{child[0].text} => {child[1].text}')
                        parametr.append(child[0].text)
                        parametr.append(child[1].text)
                        # Запись в файл
                        for nr, npar in enumerate(parametr):
                            ws.cell(row=num_row, column=num_col + nr).value = npar
                        parametr = [comp_naim, 'Computer']
                        num_row += 1

        nodes = ['CPU', 'MOBO', 'MEMORY', 'VIDEO', 'MONITOR', 'DRIVES', 'SOUND', 'NETWORK']  # Узлы
        node_par = [('Processor Name', 'Original Processor Frequency [MHz]',
                     'Number of CPU Cores', 'Number of Logical CPUs'),
                    ('Motherboard Model', 'Motherboard Chipset'),
                    ('Total Memory Size [MB]', 'Maximum Supported Memory Clock', 'Current Timing (tCAS-tRCD-tRP-tRAS)'),
                    ('Video Chipset', 'Video Chipset Codename', 'Video Memory'),
                    ('Monitor Name (Manuf)', 'Serial Number', 'Date Of Manufacture'),
                    ('Drive Controller', 'Drive Model', 'Drive Serial Number', 'Drive Capacity [MB]'), 'Audio Adapter',
                    ('Network Card', 'MAC Address', 'Drive Serial Number', 'Drive Capacity [MB]'),
                    ]

        for num_node, name_node in enumerate(nodes):
            parametr = [comp_naim]
            for elements in root.iter(name_node):  # Узлы
                for child in elements:
                    if child.tag == 'NodeName':  # Имя узла
                        child_naim = child.text
                        print(child_naim)
                        parametr.append(child_naim)
                    if child.tag == 'Property':  # Параметры узла
                        if child[0].text in node_par[num_node]:
                            print(f'  {child[0].text} => {child[1].text}')
                            parametr.append(child[0].text)
                            parametr.append(child[1].text)
                            # Запись в файл
                            for nr, npar in enumerate(parametr):
                                ws.cell(row=num_row, column=num_col + nr).value = npar
                            parametr = [comp_naim, child_naim]
                            num_row += 1

                    if child.tag == 'SubNode':  # Параметры узла
                        for sub_child in child:
                            if sub_child.tag == 'Property':  # Параметры узла
                                if sub_child[0].text in node_par[num_node]:
                                    parametr.append(sub_child[0].text)
                                    if sub_child[0].text == 'MAC Address':
                                        print(f'   {sub_child[0].text} => {sub_child[1].text.replace("-", ":")}')
                                        parametr.append(sub_child[1].text.replace("-", ":"))
                                    else:
                                        print(f'   {sub_child[0].text} => {sub_child[1].text}')
                                        parametr.append(sub_child[1].text)
                                    # Запись в файл
                                    for nr, npar in enumerate(parametr):
                                        ws.cell(row=num_row, column=num_col + nr).value = npar
                                    parametr = [comp_naim, child_naim]
                                    num_row += 1

                            if sub_child.tag == 'SubNode':
                                for par in sub_child:
                                    if par.tag == 'Property':  # Параметры узла
                                        if par[0].text in node_par[num_node]:
                                            print(f'  {par[0].text} => {par[1].text}')
                                            parametr.append(par[0].text)
                                            parametr.append(par[1].text)
                                            # Запись в файл
                                            for nr, npar in enumerate(parametr):
                                                ws.cell(row=num_row, column=num_col + nr).value = npar
                                            parametr = [comp_naim, child_naim]
                                            num_row += 1

            # Новая строка
            # num_row += 1
        # wb.save('zhelezo2020.xlsx')  # один проход цикла
    now = datetime.datetime.now()
    y = str(now.year)
    file_xls = os.path.join(BASE_DIR, "output\\zhelezo" + str(now.year) + "_" + str(now.month) + ".xlsx")
    wb.save(file_xls)
    return num_files + 1


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MyParser()

    sys.exit(app.exec_())
